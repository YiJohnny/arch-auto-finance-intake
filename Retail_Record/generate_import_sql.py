from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import csv

import openpyxl


WORKBOOK_PATH = Path("All Inventory.xlsx")
SQL_OUTPUT_PATH = Path("import_all_inventory.sql")
REVIEW_OUTPUT_PATH = Path("import_review.csv")
SUMMARY_OUTPUT_PATH = Path("import_summary.md")


STATUS_MAP = {
    "SOLD": "SOLD",
    "IN INVENTORY": "IN_INVENTORY",
    "INBOUND": "IN_BOUND",
    "TRADE IN": "IN_INVENTORY",
}

IMPORTABLE_SOURCE_STATUSES = {"SOLD", "IN INVENTORY", "INBOUND", "TRADE IN"}


@dataclass
class ImportRow:
    row_number: int
    stock_number: str
    vin: str
    year: int | None
    make: str
    model: str
    trim: str | None
    mileage: int | None
    color: str | None
    purchase_date: date
    status: str
    sale_date: date | None
    sale_price: Decimal | None
    purchase_source: str | None
    vehicle_cost: Decimal
    other_direct_cost: Decimal


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_int(value: object) -> int | None:
    text = normalize_text(value)
    if text is None:
        return None
    try:
        return int(Decimal(text))
    except Exception:
        return None


def normalize_decimal(value: object) -> Decimal:
    text = normalize_text(value)
    if text is None:
        return Decimal("0")
    return Decimal(str(value)).quantize(Decimal("0.01"))


def normalize_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    return datetime.fromisoformat(text).date()


def sql_literal(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime)):
        return f"'{value.isoformat()}'"
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def build_insert_vehicle_sql(rows: list[ImportRow]) -> list[str]:
    statements = [
        "-- Generated from All Inventory.xlsx",
        "-- Assumption: SOLD rows use Created Date as a temporary sale_date when a true sale close date is unavailable.",
        "begin;",
    ]
    for row in rows:
        statements.append(
            "\n".join(
                [
                    "insert into public.vehicles (",
                    "    stock_number,",
                    "    vin,",
                    "    year,",
                    "    make,",
                    "    model,",
                    "    trim,",
                    "    color,",
                    "    mileage,",
                    "    purchase_date,",
                    "    purchase_source,",
                    "    status,",
                    "    sale_date,",
                    "    sale_price,",
                    "    acquisition_notes",
                    ") values (",
                    f"    {sql_literal(row.stock_number)},",
                    f"    {sql_literal(row.vin)},",
                    f"    {sql_literal(row.year)},",
                    f"    {sql_literal(row.make)},",
                    f"    {sql_literal(row.model)},",
                    f"    {sql_literal(row.trim)},",
                    f"    {sql_literal(row.color)},",
                    f"    {sql_literal(row.mileage)},",
                    f"    {sql_literal(row.purchase_date)},",
                    f"    {sql_literal(row.purchase_source)},",
                    f"    {sql_literal(row.status)},",
                    f"    {sql_literal(row.sale_date)},",
                    f"    {sql_literal(row.sale_price)},",
                    f"    {sql_literal(f'Imported from All Inventory.xlsx row {row.row_number}')}",
                    ") on conflict (vin) do nothing;",
                ]
            )
        )

        if row.vehicle_cost != Decimal("0.00"):
            statements.append(
                "\n".join(
                    [
                        "insert into public.vehicle_cost_ledger (",
                        "    vehicle_id,",
                        "    entry_date,",
                        "    entry_type,",
                        "    category,",
                        "    amount,",
                        "    description,",
                        "    source_system,",
                        "    source_record_id",
                        ")",
                        "select",
                        "    v.vehicle_id,",
                        f"    {sql_literal(row.purchase_date)},",
                        "    'ACQUISITION',",
                        "    'PURCHASE',",
                        f"    {sql_literal(row.vehicle_cost)},",
                        "    'Imported initial purchase cost from legacy inventory summary',",
                        "    'LEGACY_IMPORT',",
                        f"    {sql_literal(row.vin)}",
                        "from public.vehicles v",
                        f"where v.vin = {sql_literal(row.vin)}",
                        "  and not exists (",
                        "      select 1",
                        "      from public.vehicle_cost_ledger l",
                        "      where l.vehicle_id = v.vehicle_id",
                        "        and l.entry_type = 'ACQUISITION'",
                        "        and l.category = 'PURCHASE'",
                        "  );",
                    ]
                )
            )

        if row.other_direct_cost != Decimal("0.00"):
            statements.append(
                "\n".join(
                    [
                        "insert into public.vehicle_cost_ledger (",
                        "    vehicle_id,",
                        "    entry_date,",
                        "    entry_type,",
                        "    category,",
                        "    amount,",
                        "    description,",
                        "    source_system,",
                        "    source_record_id",
                        ")",
                        "select",
                        "    v.vehicle_id,",
                        f"    {sql_literal(row.purchase_date)},",
                        "    'DIRECT_COST',",
                        "    'OTHER_DIRECT_COST',",
                        f"    {sql_literal(row.other_direct_cost)},",
                        "    'Imported historical direct costs from legacy inventory summary',",
                        "    'LEGACY_IMPORT',",
                        f"    {sql_literal(row.vin + ':other_direct_cost')}",
                        "from public.vehicles v",
                        f"where v.vin = {sql_literal(row.vin)}",
                        "  and not exists (",
                        "      select 1",
                        "      from public.vehicle_cost_ledger l",
                        "      where l.vehicle_id = v.vehicle_id",
                        "        and l.entry_type = 'DIRECT_COST'",
                        "        and l.category = 'OTHER_DIRECT_COST'",
                        "        and l.source_record_id = "
                        f"{sql_literal(row.vin + ':other_direct_cost')}",
                        "  );",
                    ]
                )
            )

    statements.append("commit;")
    return statements


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    worksheet = workbook.active
    headers = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: position for position, header in enumerate(headers)}

    imported_rows: list[ImportRow] = []
    review_rows: list[dict[str, object]] = []
    status_counter: Counter[str] = Counter()

    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue

        source_status = normalize_text(values[index["Inventory Status"]]) or "<blank>"
        status_counter[source_status] += 1

        stock_number = normalize_text(values[index["Stock Number"]])
        vin = normalize_text(values[index["VIN"]])
        make = normalize_text(values[index["Make"]])
        model = normalize_text(values[index["Model"]])
        purchase_date = normalize_date(values[index["Date In Stock"]]) or normalize_date(
            values[index["Created Date"]]
        )
        mapped_status = STATUS_MAP.get(source_status, "IN_INVENTORY")

        if source_status not in IMPORTABLE_SOURCE_STATUSES:
            review_rows.append(
                {
                    "row_number": row_number,
                    "stock_number": stock_number or "",
                    "vin": vin or "",
                    "source_status": source_status,
                    "reason": "excluded_source_status",
                }
            )
            continue

        if not vin or not stock_number or not make or not model or not purchase_date:
            review_rows.append(
                {
                    "row_number": row_number,
                    "stock_number": stock_number or "",
                    "vin": vin or "",
                    "source_status": source_status,
                    "reason": "missing_required_field",
                }
            )
            continue

        selling_price = normalize_text(values[index["Selling Price"]])
        created_date = normalize_date(values[index["Created Date"]])
        sale_price = normalize_decimal(selling_price) if selling_price is not None else None
        sale_date = created_date if mapped_status in {"SOLD", "WHOLESALE"} else None

        if mapped_status in {"SOLD", "WHOLESALE"} and sale_price is None:
            review_rows.append(
                {
                    "row_number": row_number,
                    "stock_number": stock_number,
                    "vin": vin,
                    "source_status": source_status,
                    "reason": "missing_sale_price_for_sold_status",
                }
            )
            continue

        vehicle_cost = normalize_decimal(values[index["Vehicle Cost"]])
        total_cost = normalize_decimal(values[index["Total Cost"]])
        other_direct_cost = (total_cost - vehicle_cost).quantize(Decimal("0.01"))

        imported_rows.append(
            ImportRow(
                row_number=row_number,
                stock_number=stock_number,
                vin=vin,
                year=normalize_int(values[index["Year"]]),
                make=make,
                model=model,
                trim=normalize_text(values[index["Trim"]]),
                mileage=normalize_int(values[index["Odometer"]]),
                color=normalize_text(values[index["Exterior Color"]]),
                purchase_date=purchase_date,
                status=mapped_status,
                sale_date=sale_date,
                sale_price=sale_price if mapped_status in {"SOLD", "WHOLESALE"} else None,
                purchase_source="Trade-In" if source_status == "TRADE IN" else source_status,
                vehicle_cost=vehicle_cost,
                other_direct_cost=other_direct_cost,
            )
        )

    SQL_OUTPUT_PATH.write_text("\n\n".join(build_insert_vehicle_sql(imported_rows)) + "\n")

    with REVIEW_OUTPUT_PATH.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["row_number", "stock_number", "vin", "source_status", "reason"],
        )
        writer.writeheader()
        writer.writerows(review_rows)

    summary_lines = [
        "# Import Summary",
        "",
        f"- Source rows scanned: `{sum(status_counter.values())}`",
        f"- Rows prepared for import: `{len(imported_rows)}`",
        f"- Rows sent to manual review: `{len(review_rows)}`",
        "",
        "## Source Status Counts",
        "",
    ]
    for status, count in status_counter.most_common():
        summary_lines.append(f"- `{status}`: `{count}`")

    summary_lines.extend(
        [
            "",
            "## Import Assumptions",
            "",
            "- `APPRAISAL` rows are excluded from import by default.",
            "- `IN INVENTORY` is mapped to `IN_INVENTORY`.",
            "- `TRADE IN` is mapped to `IN_INVENTORY` with `purchase_source = 'Trade-In'`.",
            "- `SOLD` rows use `Created Date` as a temporary `sale_date` because the export does not contain an explicit sold date.",
            "- `SOLD` rows missing `Selling Price` are excluded and written to `import_review.csv`.",
            "- Rows missing `Stock Number`, `VIN`, `Make`, `Model`, or `purchase_date` are excluded and written to `import_review.csv`.",
        ]
    )
    SUMMARY_OUTPUT_PATH.write_text("\n".join(summary_lines) + "\n")


if __name__ == "__main__":
    main()
